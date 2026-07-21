#!/usr/bin/env python3
"""Genera archivos .xlsx sintéticos de un tamaño objetivo (en MB) y mide
cuánto tarda `termsheet.io.xlsx_io.load_workbook` en abrirlos y cuánta
memoria RSS pico usa.

La apertura se ejecuta en un subproceso aparte con un límite de memoria
(RLIMIT_AS) para no arriesgar al resto de servicios de la máquina si un
archivo enorme hace que el proceso se descontrole: en vez de dejar que el
OOM-killer del sistema mate un proceso cualquiera (podría ser cualquier
otro servicio del servidor), el propio subproceso recibe un MemoryError
limpio y lo reportamos como fallo controlado.

Uso:
    python scripts/benchmark_large_xlsx.py --sizes 100 500 1000
    python scripts/benchmark_large_xlsx.py --sizes 100 --keep-files --outdir /tmp/bench
"""

from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
import time
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent

# 20 columnas por fila: mezcla de texto, enteros, decimales, fecha y una
# fórmula que referencia otra columna — para que el archivo generado se
# parezca a una hoja de cálculo real y no solo a relleno repetitivo.
N_COLS = 20


def _row(i: int) -> list:
    from datetime import date, timedelta

    base_date = date(2024, 1, 1)
    return [
        f"Producto {i:07d}",
        f"SKU-{i % 10000:05d}",
        i,
        round(i * 1.3457, 2),
        (base_date + timedelta(days=i % 3650)).isoformat(),
        f"=D{i + 2}*1.21",  # IVA sobre la columna anterior, como en una hoja real
        "activo" if i % 3 else "descatalogado",
        i % 5,
        round((i % 97) / 3.0, 3),
        f"nota de texto libre para la fila {i} " * 2,
    ] * 2  # repetido para llegar a N_COLS=20 sin complicar la lista


def _calibrate_bytes_per_row(sample_rows: int = 20_000) -> float:
    """Genera una muestra pequeña para estimar cuántas filas hacen falta
    para alcanzar el tamaño objetivo, ya que con Workbook(write_only=True)
    no se puede consultar el tamaño real del archivo hasta hacer save()."""
    import tempfile

    from openpyxl import Workbook

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        sample_path = tmp.name
    try:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Datos")
        ws.append([f"col{i}" for i in range(N_COLS)])
        for i in range(sample_rows):
            ws.append(_row(i))
        wb.save(sample_path)
        size = os.path.getsize(sample_path)
        return size / sample_rows
    finally:
        os.unlink(sample_path)


def generate_xlsx(path: Path, target_mb: int) -> None:
    from openpyxl import Workbook

    target_bytes = target_mb * 1024 * 1024
    bytes_per_row = _calibrate_bytes_per_row()
    n_rows = max(1, int(target_bytes / bytes_per_row))

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Datos")
    ws.append([f"col{i}" for i in range(N_COLS)])
    for i in range(n_rows):
        ws.append(_row(i))
    wb.save(str(path))


def _child_main(path: str, mem_cap_mb: int) -> None:
    """Se ejecuta en el subproceso hijo: aplica el límite de memoria, abre
    el archivo con el loader real de termsheet, y escribe el resultado
    (tiempo + memoria pico) como JSON en stdout."""
    import resource

    sys.path.insert(0, str(REPO_ROOT))

    if mem_cap_mb > 0:
        cap_bytes = mem_cap_mb * 1024 * 1024
        try:
            resource.setrlimit(resource.RLIMIT_AS, (cap_bytes, cap_bytes))
        except (ValueError, OSError):
            pass  # algunas plataformas (p.ej. macOS) restringen bajar este límite

    result = {"success": False, "elapsed_seconds": None, "peak_rss_mb": None, "error": None}
    try:
        from termsheet.io.xlsx_io import load_workbook

        start = time.perf_counter()
        wb = load_workbook(path)
        n_sheets = len(wb.sheets)
        elapsed = time.perf_counter() - start
        result["success"] = True
        result["elapsed_seconds"] = round(elapsed, 2)
        result["n_sheets"] = n_sheets
    except MemoryError as exc:
        result["error"] = f"MemoryError: superó el límite de {mem_cap_mb} MB"
    except Exception as exc:  # noqa: BLE001 — queremos capturar y reportar cualquier fallo
        result["error"] = f"{type(exc).__name__}: {exc}"
    finally:
        usage = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss
        # Linux reporta ru_maxrss en KB; macOS en bytes.
        peak_mb = usage / 1024 if sys.platform != "darwin" else usage / (1024 * 1024)
        result["peak_rss_mb"] = round(peak_mb, 1)

    print(json.dumps(result))


def benchmark_open(path: Path, mem_cap_mb: int, timeout_seconds: int = 900) -> dict:
    """Lanza el subproceso hijo que hace la apertura real y devuelve su
    resultado. Si el proceso muere sin más explicación (p.ej. el kernel lo
    mata directamente por presión de memoria, esquivando incluso el límite
    de RLIMIT_AS), se reporta como fallo en vez de dejar que la excepción
    se propague y tumbe el resto del benchmark."""
    cmd = [sys.executable, str(Path(__file__).resolve()), "--child", str(path), "--mem-cap-mb", str(mem_cap_mb)]
    try:
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout_seconds)
    except subprocess.TimeoutExpired:
        return {"success": False, "elapsed_seconds": None, "peak_rss_mb": None, "error": f"timeout tras {timeout_seconds}s"}

    if proc.returncode != 0 or not proc.stdout.strip():
        return {
            "success": False,
            "elapsed_seconds": None,
            "peak_rss_mb": None,
            "error": f"el subproceso terminó con código {proc.returncode} (probablemente matado por el kernel "
            f"por falta de memoria real, ya que esquiva incluso el límite RLIMIT_AS). stderr: {proc.stderr[-500:]}",
        }
    try:
        return json.loads(proc.stdout.strip().splitlines()[-1])
    except json.JSONDecodeError:
        return {"success": False, "elapsed_seconds": None, "peak_rss_mb": None, "error": f"salida no-JSON: {proc.stdout[-500:]}"}


def _free_memory_mb() -> int:
    try:
        with open("/proc/meminfo") as fh:
            for line in fh:
                if line.startswith("MemAvailable:"):
                    return int(line.split()[1]) // 1024
    except FileNotFoundError:
        pass
    return -1  # no disponible (p.ej. macOS/Windows) — sin dato de protección


def run_benchmarks(sizes_mb: list[int], outdir: Path, keep_files: bool) -> list[dict]:
    outdir.mkdir(parents=True, exist_ok=True)
    results = []
    for size_mb in sizes_mb:
        path = outdir / f"bench_{size_mb}mb.xlsx"
        print(f"\n== {size_mb} MB ==")

        free_mb = _free_memory_mb()
        # Tope de seguridad: no más del 70% de la memoria libre real en este
        # momento, así un benchmark que se dispare de memoria falla limpio
        # en su propio subproceso en vez de arriesgar al resto de servicios
        # de este servidor compartido.
        mem_cap_mb = max(256, int(free_mb * 0.7)) if free_mb > 0 else 4096
        print(f"Memoria libre actual: {free_mb} MB -> límite aplicado al subproceso: {mem_cap_mb} MB")

        print(f"Generando {path} (~{size_mb} MB)...")
        t0 = time.perf_counter()
        generate_xlsx(path, size_mb)
        gen_elapsed = time.perf_counter() - t0
        real_size_mb = path.stat().st_size / (1024 * 1024)
        print(f"Generado en {gen_elapsed:.1f}s, tamaño real {real_size_mb:.1f} MB")

        print("Abriendo y midiendo...")
        result = benchmark_open(path, mem_cap_mb)
        result["target_mb"] = size_mb
        result["real_file_mb"] = round(real_size_mb, 1)
        results.append(result)

        if result["success"]:
            print(f"OK — {result['elapsed_seconds']}s, pico de memoria {result['peak_rss_mb']} MB")
        else:
            print(f"FALLO — {result['error']}")

        if not keep_files:
            path.unlink(missing_ok=True)

    return results


def print_markdown_table(results: list[dict]) -> None:
    print("\n| Tamaño archivo | Tiempo de apertura | Memoria pico (RSS) | Resultado |")
    print("|---|---|---|---|")
    for r in results:
        size = f"{r['real_file_mb']} MB"
        if r["success"]:
            print(f"| {size} | {r['elapsed_seconds']} s | {r['peak_rss_mb']} MB | OK |")
        else:
            print(f"| {size} | — | {r.get('peak_rss_mb', '—')} MB | Fallo: {r['error']} |")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--sizes", type=int, nargs="+", default=[100, 500, 1000], help="Tamaños objetivo en MB")
    parser.add_argument("--outdir", type=Path, default=Path("/tmp/termsheet_bench"))
    parser.add_argument("--keep-files", action="store_true", help="No borrar los .xlsx generados al terminar")
    parser.add_argument("--child", type=str, default=None, help=argparse.SUPPRESS)
    parser.add_argument("--mem-cap-mb", type=int, default=0, help=argparse.SUPPRESS)
    args = parser.parse_args()

    if args.child is not None:
        _child_main(args.child, args.mem_cap_mb)
        return

    results = run_benchmarks(args.sizes, args.outdir, args.keep_files)
    print_markdown_table(results)


if __name__ == "__main__":
    main()
