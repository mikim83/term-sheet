"""Carga y guarda libros en formato .xlsx real usando openpyxl.

Compatible con archivos de Excel y con .xlsx exportados manualmente desde
Google Sheets (Archivo -> Descargar -> Microsoft Excel).
"""

from __future__ import annotations

import datetime
import xml.etree.ElementTree as ET

from openpyxl import Workbook as XlWorkbook
from openpyxl import load_workbook as xl_load_workbook
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..model.cell import CellFormat
from ..model.formatting import FORMATS, XLSX_CODE_TO_KEY
from ..model.workbook import Sheet, Workbook


def _rgb_from_xl_color(color) -> str | None:
    """Extrae "RRGGBB" de un openpyxl.styles.colors.Color, o None si no tiene
    un valor RGB directo (p.ej. colores de tema, que no mapean 1:1 a hex)."""
    if color is None or not getattr(color, "rgb", None):
        return None
    rgb = color.rgb
    if not isinstance(rgb, str) or len(rgb) < 6:
        return None
    return rgb[-6:]  # los colores de Excel vienen en ARGB (8 hex); nos quedamos con RGB


def _border_style_and_color(xl_cell) -> tuple[str | None, str]:
    """Se guarda un único grosor/estilo de línea por celda (no por lado):
    cogemos el primer lado con borde definido (top/right/bottom/left, en ese
    orden) como representativo. Es una simplificación deliberada — ver
    CellFormat.border_style — pero evita perder por completo bordes reales
    de un .xlsx al reabrirlo y volver a guardarlo."""
    border = xl_cell.border
    if border is None:
        return None, "808080"
    for side in (border.top, border.right, border.bottom, border.left):
        if side is not None and side.style:
            color = _rgb_from_xl_color(side.color) or "808080"
            return side.style, color
    return None, "808080"


def _read_column_widths_streaming(xl_wb, xl_sheet) -> dict[int, int]:
    """Los anchos de columna (`<cols><col .../></cols>`) no están accesibles
    vía `column_dimensions` en modo `read_only` de openpyxl (ver
    `load_workbook` de más abajo, y por qué usamos ese modo). Como ese bloque
    siempre aparece ANTES de `<sheetData>` (que es lo que realmente pesa en
    archivos grandes), lo leemos con un parseo streaming aparte que se
    detiene en cuanto empieza `<sheetData>` — así el coste sigue siendo
    ínfimo (unos KB) sin importar lo grande que sea la hoja."""
    widths: dict[int, int] = {}
    try:
        stream = xl_wb._archive.open(xl_sheet._worksheet_path)
    except (KeyError, AttributeError):
        return widths
    try:
        for event, elem in ET.iterparse(stream, events=("start", "end")):
            tag = elem.tag.rsplit("}", 1)[-1]
            if tag == "col" and event == "end":
                width = elem.get("width")
                col_min = elem.get("min")
                col_max = elem.get("max")
                if width and col_min and col_max:
                    for col in range(int(col_min), int(col_max) + 1):
                        widths[col] = int(float(width))
            elif tag == "sheetData" and event == "start":
                break
    finally:
        stream.close()
    return widths


def load_workbook(path: str) -> Workbook:
    # read_only=True usa el parser en streaming de openpyxl en vez de
    # construir un grafo de objetos completo con cada celda del archivo —
    # medido: reduce la memoria pico de abrir un .xlsx grande más de 20x
    # (ver README, sección Benchmark). El acceso a valor/estilo por celda
    # (font, fill, border, number_format) funciona igual en ambos modos.
    xl_wb = xl_load_workbook(path, data_only=False, read_only=True)
    wb = Workbook(sheets=[])
    for xl_sheet in xl_wb.worksheets:
        sheet = Sheet(xl_sheet.title)
        col_widths = _read_column_widths_streaming(xl_wb, xl_sheet)
        for row in xl_sheet.iter_rows():
            for xl_cell in row:
                if xl_cell.value is None:
                    continue
                raw = _to_raw(xl_cell.value)
                cell = sheet.set_raw(xl_cell.row, xl_cell.column, raw)
                xl_format = xl_cell.number_format
                if isinstance(xl_cell.value, (datetime.date, datetime.datetime)):
                    # Excel usa muchos códigos de formato de fecha distintos (d/MM/yyyy,
                    # dd-mm-yy, etc.) que no coinciden con nuestro FORMATS["date_dmy"].xlsx_code
                    # — si el valor ya es una fecha, siempre usamos nuestro único formato de fecha.
                    number_format = "date_dmy"
                else:
                    number_format = XLSX_CODE_TO_KEY.get(xl_format, xl_format if xl_format != "General" else None)
                font_color = _rgb_from_xl_color(xl_cell.font.color) if xl_cell.font else None
                bg_color = None
                if xl_cell.fill and xl_cell.fill.patternType == "solid":
                    bg_color = _rgb_from_xl_color(xl_cell.fill.fgColor)
                border_style, border_color = _border_style_and_color(xl_cell)
                cell.fmt = CellFormat(
                    bold=bool(xl_cell.font and xl_cell.font.bold),
                    italic=bool(xl_cell.font and xl_cell.font.italic),
                    align=(xl_cell.alignment.horizontal or "left") if xl_cell.alignment else "left",
                    number_format=number_format,
                    font_color=font_color,
                    bg_color=bg_color,
                    border_style=border_style,
                    border_color=border_color,
                )
        sheet.col_widths.update(col_widths)
        wb.sheets.append(sheet)
    if not wb.sheets:
        wb.sheets.append(Sheet("Hoja1"))
    wb.path = path
    return wb


def _to_raw(value) -> str:
    if isinstance(value, str) and value.startswith("="):
        return value
    if isinstance(value, datetime.datetime):
        if value.time() == datetime.time():
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    return str(value)


def save_workbook(workbook: Workbook, path: str) -> None:
    xl_wb = XlWorkbook()
    xl_wb.remove(xl_wb.active)
    for sheet in workbook.sheets:
        xl_sheet = xl_wb.create_sheet(title=sheet.name)
        for row, col, cell in sheet.iter_cells():
            xl_cell = xl_sheet.cell(row=row, column=col)
            xl_cell.value = _cell_value_for_xlsx(cell)
            if cell.fmt.bold or cell.fmt.italic or cell.fmt.font_color:
                font_kwargs = {"bold": cell.fmt.bold, "italic": cell.fmt.italic}
                if cell.fmt.font_color:
                    font_kwargs["color"] = cell.fmt.font_color
                xl_cell.font = xl_cell.font.copy(**font_kwargs)
            if cell.fmt.align != "left":
                xl_cell.alignment = xl_cell.alignment.copy(horizontal=cell.fmt.align)
            if cell.fmt.number_format:
                known = FORMATS.get(cell.fmt.number_format)
                xl_cell.number_format = known.xlsx_code if known else cell.fmt.number_format
            if cell.fmt.bg_color:
                xl_cell.fill = PatternFill(fgColor=cell.fmt.bg_color, fill_type="solid")
            if cell.fmt.border_style:
                side = Side(style=cell.fmt.border_style, color=cell.fmt.border_color)
                xl_cell.border = Border(left=side, right=side, top=side, bottom=side)
        for col, width in sheet.col_widths.items():
            xl_sheet.column_dimensions[get_column_letter(col)].width = width
    xl_wb.save(path)
    workbook.path = path


def _cell_value_for_xlsx(cell):
    if cell.is_formula:
        return cell.raw
    raw = cell.raw
    try:
        if "." in raw:
            return float(raw)
        return int(raw)
    except ValueError:
        return raw
