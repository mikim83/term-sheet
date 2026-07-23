import tempfile
from pathlib import Path

from termsheet.io.xlsx_io import load_workbook, save_workbook
from termsheet.model.formula_engine import Engine
from termsheet.model.workbook import Workbook


def test_roundtrip_preserves_formulas_and_multi_sheet():
    wb = Workbook(sheets=[])
    s1 = wb.add_sheet("Datos")
    s1.set_raw(1, 1, "5")
    s1.set_raw(1, 2, "=A1*2")
    s2 = wb.add_sheet("Resumen")
    s2.set_raw(1, 1, "hola")

    with tempfile.TemporaryDirectory() as tmp:
        path = str(Path(tmp) / "test.xlsx")
        save_workbook(wb, path)

        wb2 = load_workbook(path)
        assert [s.name for s in wb2.sheets] == ["Datos", "Resumen"]
        datos = wb2.sheet_by_name("Datos")
        assert datos.get_raw(1, 1) == "5"
        assert datos.get_raw(1, 2) == "=A1*2"

        eng = Engine(wb2)
        eng.recalculate()
        assert datos.get_cell(1, 2).value == 10


def test_roundtrip_preserves_currency_and_date_format():
    wb = Workbook(sheets=[])
    sheet = wb.add_sheet("Datos")
    sheet.set_raw(1, 1, "1234.5")
    sheet.get_cell(1, 1).fmt.number_format = "eur_es"
    sheet.set_raw(2, 1, "2026-07-13")
    sheet.get_cell(2, 1).fmt.number_format = "date_dmy"

    with tempfile.TemporaryDirectory() as tmp:
        path = str(Path(tmp) / "formatos.xlsx")
        save_workbook(wb, path)

        wb2 = load_workbook(path)
        datos = wb2.sheet_by_name("Datos")
        Engine(wb2).recalculate()
        assert datos.get_cell(1, 1).fmt.number_format == "eur_es"
        assert datos.get_cell(1, 1).display() == "1.234,50 €"
        assert datos.get_cell(2, 1).fmt.number_format == "date_dmy"
        assert datos.get_cell(2, 1).display() == "13/07/2026"


def test_roundtrip_preserves_cross_sheet_formula():
    wb = Workbook(sheets=[])
    s1 = wb.add_sheet("Datos")
    s2 = wb.add_sheet("Resumen")
    s1.set_raw(1, 1, "100")
    s2.set_raw(1, 1, "=Datos!A1*2")

    with tempfile.TemporaryDirectory() as tmp:
        path = str(Path(tmp) / "cross_sheet.xlsx")
        save_workbook(wb, path)

        wb2 = load_workbook(path)
        resumen = wb2.sheet_by_name("Resumen")
        assert resumen.get_raw(1, 1) == "=Datos!A1*2"

        Engine(wb2).recalculate()
        assert resumen.get_cell(1, 1).value == 200


def test_roundtrip_preserves_font_color_bg_color_and_border():
    wb = Workbook(sheets=[])
    s1 = wb.add_sheet("Datos")
    styled = s1.set_raw(1, 1, "hola")
    styled.fmt.font_color = "FF0000"
    styled.fmt.bg_color = "FFFF00"
    styled.fmt.border_style = "thick"
    styled.fmt.border_color = "000000"
    s1.set_raw(2, 1, "sin estilo propio")  # se queda con el default (dashed gris)

    with tempfile.TemporaryDirectory() as tmp:
        path = str(Path(tmp) / "styled.xlsx")
        save_workbook(wb, path)

        wb2 = load_workbook(path)
        datos = wb2.sheet_by_name("Datos")
        styled2 = datos.get_cell(1, 1)
        assert styled2.fmt.font_color == "FF0000"
        assert styled2.fmt.bg_color == "FFFF00"
        assert styled2.fmt.border_style == "thick"
        assert styled2.fmt.border_color == "000000"

        default2 = datos.get_cell(2, 1)
        assert default2.fmt.border_style == "dashed"
        assert default2.fmt.border_color == "808080"
        assert default2.fmt.font_color is None
        assert default2.fmt.bg_color is None


def test_load_workbook_reads_colors_from_a_real_excel_file():
    """Compatibilidad con un .xlsx creado directamente con openpyxl (como lo
    haría Excel real), no generado por nuestro propio save_workbook."""
    from openpyxl import Workbook as XlWorkbook
    from openpyxl.styles import Border, Font, PatternFill, Side

    xl_wb = XlWorkbook()
    ws = xl_wb.active
    ws.title = "Hoja1"
    ws["A1"] = "importado"
    ws["A1"].font = Font(color="FF3366CC")
    ws["A1"].fill = PatternFill(fgColor="FFEEEE00", fill_type="solid")
    side = Side(style="dotted", color="FF444444")
    ws["A1"].border = Border(left=side, right=side, top=side, bottom=side)
    ws["B1"] = "sin estilo"

    with tempfile.TemporaryDirectory() as tmp:
        path = str(Path(tmp) / "external.xlsx")
        xl_wb.save(path)

        wb = load_workbook(path)
        sheet = wb.sheet_by_name("Hoja1")
        a1 = sheet.get_cell(1, 1)
        assert a1.fmt.font_color == "3366CC"
        assert a1.fmt.bg_color == "EEEE00"
        assert a1.fmt.border_style == "dotted"
        assert a1.fmt.border_color == "444444"

        b1 = sheet.get_cell(1, 2)
        assert b1.fmt.font_color is None
        assert b1.fmt.bg_color is None
        assert b1.fmt.border_style is None  # el archivo externo no le puso borde


def test_load_workbook_does_not_crash_on_gradient_fill():
    """Regresión: un archivo real (p.ej. plantillas descargadas) puede tener
    celdas con relleno degradado (GradientFill), que a diferencia de
    PatternFill no tiene el atributo `patternType` — cargarlo no debe reventar,
    simplemente no se importa un color de fondo plano para esa celda."""
    from openpyxl import Workbook as XlWorkbook
    from openpyxl.styles import GradientFill

    xl_wb = XlWorkbook()
    ws = xl_wb.active
    ws.title = "Hoja1"
    ws["A1"] = "con degradado"
    ws["A1"].fill = GradientFill(stop=("FFFFFFFF", "FF000000"))
    ws["B1"] = "normal"

    with tempfile.TemporaryDirectory() as tmp:
        path = str(Path(tmp) / "gradient.xlsx")
        xl_wb.save(path)

        wb = load_workbook(path)  # no debe lanzar AttributeError
        sheet = wb.sheet_by_name("Hoja1")
        assert sheet.get_cell(1, 1).fmt.bg_color is None
        assert sheet.get_cell(1, 2).fmt.bg_color is None
